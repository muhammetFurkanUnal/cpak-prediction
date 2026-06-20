#!/usr/bin/env python3
"""
Write model measurements (metrics.csv) into metrics.xlsx as side-by-side columns
================================================================================

The ground-truth block (columns A–I) is left untouched. The model's measurements
are added in a parallel block starting at column K, row-aligned with the
ground truth, so the sheet shows hand-measured vs model values side by side.

Mirroring how the ground-truth block works, only the two raw angles are written
as values — MPTA (col K) and LDFA (col L). Everything derived (MPTA-LDFA,
MPTA+LDFA, DİZİLİM, EKLEM, SINIFLAMA) is an Excel formula referencing K/L, i.e.
the exact same formulas the A–I block uses on C/D.

Only cpak rows (GRAFİ = 1) carry model values; kneeAP rows and cpak rows whose
image is missing from the model output stay blank (no values, no formulas).

Usage
-----
    .venv/bin/python scripts/write_model_to_xlsx.py \\
        --csv data/metrics.csv --xlsx data/metrics.xlsx
"""

import argparse
import csv
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# model block: K..Q (1-based 11..17), with J (10) left blank as a separator.
COL_MPTA = 11   # K
COL_LDFA = 12   # L
COL_DIFF = 13   # M
COL_SUM = 14    # N
COL_ALIGN = 15  # O
COL_JOINT = 16  # P
COL_TYPE = 17   # Q

HEADERS = {
    COL_MPTA: "MPTA (model)",
    COL_LDFA: "LDFA (model)",
    COL_DIFF: "MPTA-LDFA (model)",
    COL_SUM: "MPTA+LDFA (model)",
    COL_ALIGN: "DİZİLİM (model)",
    COL_JOINT: "EKLEM (model)",
    COL_TYPE: "SINIFLAMA (model)",
}


def derived_formulas(r: int) -> dict:
    """The same formulas the A–I block uses, retargeted onto the model's K/L."""
    return {
        COL_DIFF: f"=K{r}-L{r}",
        COL_SUM: f"=K{r}+L{r}",
        COL_ALIGN: f'=IF(AND(M{r}>=-2,M{r}<=2),"nötr",IF(M{r}<-2,"varus","valgus"))',
        COL_JOINT: f'=IF(AND(N{r}>=177,N{r}<=183),"nötr",IF(N{r}<177,"apex distal","apex proksimal"))',
        COL_TYPE: (
            f'=IFS(AND(O{r}="varus", P{r}="apex distal"), "Tip 1", '
            f'AND(O{r}="nötr", P{r}="apex distal"), "Tip 2", '
            f'AND(O{r}="valgus", P{r}="apex distal"), "Tip 3", '
            f'AND(O{r}="varus", P{r}="nötr"), "Tip 4", '
            f'AND(O{r}="nötr", P{r}="nötr"), "Tip 5", '
            f'AND(O{r}="valgus", P{r}="nötr"), "Tip 6", '
            f'AND(O{r}="varus", P{r}="apex proksimal"), "Tip 7", '
            f'AND(O{r}="nötr", P{r}="apex proksimal"), "Tip 8", '
            f'AND(O{r}="valgus", P{r}="apex proksimal"), "Tip 9", '
            f'TRUE, "")'
        ),
    }


def norm(s: str) -> str:
    return (s or "").strip().lower()


def main(argv=None) -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", type=Path, default=PROJECT_ROOT / "data" / "metrics.csv")
    parser.add_argument("--xlsx", type=Path, default=PROJECT_ROOT / "data" / "metrics.xlsx")
    args = parser.parse_args(argv)

    for p in (args.csv, args.xlsx):
        if not p.exists():
            sys.exit(f"Required file not found: {p}")

    with open(args.csv, newline="", encoding="utf-8") as f:
        csv_rows = list(csv.reader(f))[1:]  # drop header

    wb = openpyxl.load_workbook(args.xlsx)  # keep formulas + formatting in A–I
    ws = wb.active

    # headers in row 1
    for col, text in HEADERS.items():
        ws.cell(row=1, column=col, value=text)

    filled = blank = mismatch = 0
    for i, csv_row in enumerate(csv_rows):
        r = i + 2  # xlsx row (header is row 1)
        # sanity: HASTA/GRAFİ must line up between csv and xlsx
        x_hasta = ws.cell(row=r, column=1).value
        x_grafi = ws.cell(row=r, column=2).value
        x_grafi = str(int(x_grafi)) if isinstance(x_grafi, float) else str(x_grafi or "").strip()
        if norm(str(x_hasta)) != norm(csv_row[0]) or x_grafi != csv_row[1].strip():
            mismatch += 1
            print(f"  ! row {r}: csv=({csv_row[0]!r},{csv_row[1]!r}) "
                  f"xlsx=({x_hasta!r},{x_grafi!r}) — skipped", file=sys.stderr)
            continue

        mpta, ldfa = csv_row[2].strip(), csv_row[3].strip()
        if mpta == "" or ldfa == "":   # kneeAP / no-image -> stays blank
            blank += 1
            continue

        ws.cell(row=r, column=COL_MPTA, value=float(mpta))
        ws.cell(row=r, column=COL_LDFA, value=float(ldfa))
        for col, formula in derived_formulas(r).items():
            ws.cell(row=r, column=col, value=formula)
        filled += 1

    if mismatch:
        sys.exit(f"\nAborted before saving: {mismatch} row(s) did not line up. "
                 f"xlsx left unchanged.")

    wb.save(args.xlsx)
    print(f"Saved: {args.xlsx.relative_to(PROJECT_ROOT)}")
    print(f"  model rows filled (K–Q): {filled}")
    print(f"  left blank             : {blank}  (kneeAP / no image)")


if __name__ == "__main__":
    main()
